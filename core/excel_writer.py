"""
Tahap 6-8: menulis laporan Excel profesional.

- Sheet "Main Data" + sheet "Summary".
- Conditional Formatting otomatis pada kolom Harga Main Data
  (hijau/kuning/merah) berbanding Harga Marketplace Termurah.
- Format Rupiah, Auto Filter, Freeze Pane, Auto Fit lebar kolom.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from . import config
from .logger import get_logger

log = get_logger()

RUPIAH_FMT = '"Rp" #,##0'
PERCENT_FMT = '0.00"%"'

# Kolom yang diformat sebagai Rupiah.
_MONEY_COLUMNS = {
    config.MAIN_PRICE_COLUMN,
    "Harga Marketplace Termurah",
    "Harga Marketplace Rata-rata",
    "Selisih Harga",
}
_PERCENT_COLUMNS = {"Persentase Selisih"}


def _col_letter(df: pd.DataFrame, name: str) -> str | None:
    if name not in df.columns:
        return None
    return get_column_letter(df.columns.get_loc(name) + 1)


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor=config.COLOR_HEADER)
    font = Font(color=config.COLOR_HEADER_TEXT, bold=True, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 28


def _apply_number_formats(ws, df: pd.DataFrame, nrows: int) -> None:
    for name in _MONEY_COLUMNS:
        letter = _col_letter(df, name)
        if not letter:
            continue
        for r in range(2, nrows + 2):
            ws[f"{letter}{r}"].number_format = RUPIAH_FMT
    for name in _PERCENT_COLUMNS:
        letter = _col_letter(df, name)
        if not letter:
            continue
        for r in range(2, nrows + 2):
            ws[f"{letter}{r}"].number_format = PERCENT_FMT


def _apply_conditional_formatting(ws, df: pd.DataFrame, nrows: int) -> None:
    """
    Warnai kolom Harga Main Data otomatis vs Harga Marketplace Termurah.

    Hijau : harga kita < termurah  (lebih murah)
    Kuning: harga kita = termurah  (sama)
    Merah : harga kita > termurah  (lebih mahal)
    Tanpa data marketplace -> tidak diwarnai.
    """
    main = _col_letter(df, config.MAIN_PRICE_COLUMN)
    term = _col_letter(df, "Harga Marketplace Termurah")
    if not main or not term or nrows == 0:
        return

    rng = f"{main}2:{main}{nrows + 1}"
    # Referensi kolom absolut ($X), baris relatif (2) -> menyesuaikan tiap baris.
    guard = f'${main}2<>"",${term}2<>""'

    rules = [
        (f"AND({guard},${main}2<${term}2)", config.COLOR_GREEN, config.COLOR_GREEN_TEXT),
        (f"AND({guard},${main}2=${term}2)", config.COLOR_YELLOW, config.COLOR_YELLOW_TEXT),
        (f"AND({guard},${main}2>${term}2)", config.COLOR_RED, config.COLOR_RED_TEXT),
    ]
    for formula, fill_color, font_color in rules:
        ws.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[formula],
                fill=PatternFill("solid", fgColor=fill_color),
                font=Font(color=font_color),
                stopIfTrue=True,
            ),
        )


def _apply_row_highlight(ws, df: pd.DataFrame, nrows: int, ncols: int) -> None:
    """
    Warnai SELURUH baris data berdasarkan Status Harga (Perubahan 2).

    Berbeda dari conditional formatting di atas (yang hidup mengikuti rumus dan
    hanya menyasar kolom harga), fungsi ini menuliskan PatternFill statis lewat
    openpyxl sehingga warna benar-benar tersimpan di dalam file .xlsx:

      Hijau   : Harga Main Data < Harga Marketplace  (lebih murah)
      Kuning  : Harga Main Data = Harga Marketplace  (sama)
      Merah   : Harga Main Data > Harga Marketplace  (lebih mahal)
      Abu-abu : produk tidak ditemukan / harga marketplace kosong
    """
    if "Status Harga" not in df.columns or nrows == 0:
        return

    grey = PatternFill("solid", fgColor=config.COLOR_GREY)
    fills = {
        config.STATUS_CHEAPER: PatternFill("solid", fgColor=config.COLOR_GREEN),
        config.STATUS_EQUAL: PatternFill("solid", fgColor=config.COLOR_YELLOW),
        config.STATUS_MORE_EXPENSIVE: PatternFill("solid", fgColor=config.COLOR_RED),
        config.STATUS_NO_DATA: grey,
    }

    statuses = df["Status Harga"].tolist()
    counts: dict[str, int] = {}
    for i, status in enumerate(statuses):
        # Status tak dikenal / kosong diperlakukan sebagai tanpa data -> abu-abu.
        fill = fills.get(status, grey)
        row = i + 2  # baris 1 = header
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = fill
        key = status if status in fills else config.STATUS_NO_DATA
        counts[key] = counts.get(key, 0) + 1

    log.info("  Highlight baris: %s", counts)


def _autofit(ws, df: pd.DataFrame) -> None:
    """Perkirakan lebar kolom dari isi (dibatasi agar tidak terlalu lebar)."""
    sample = df.head(1000)
    for i, col in enumerate(df.columns, start=1):
        header_len = len(str(col))
        try:
            body_len = int(sample[col].astype(str).str.len().max() or 0)
        except Exception:  # noqa: BLE001
            body_len = 0
        width = min(max(header_len, body_len) + 2, 55)
        ws.column_dimensions[get_column_letter(i)].width = max(width, 10)


def _write_summary_sheet(wb, summary: dict) -> None:
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    title = ws.cell(row=1, column=1, value="Ringkasan Analisa Harga")
    title.font = Font(bold=True, size=14, color=config.COLOR_HEADER)
    ws.merge_cells("A1:B1")

    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_keys = {"Rata-rata Selisih Harga"}

    r = 3
    for key, val in summary.items():
        kc = ws.cell(row=r, column=1, value=key)
        vc = ws.cell(row=r, column=2, value=val)
        kc.font = Font(bold=True)
        kc.fill = PatternFill("solid", fgColor="FFF2F2F2")
        kc.border = border
        vc.border = border
        if key in money_keys and isinstance(val, (int, float)):
            vc.number_format = RUPIAH_FMT
        r += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    # Legenda warna status.
    r += 1
    ws.cell(row=r, column=1, value="Legenda Warna").font = Font(bold=True)
    legend = [
        ("Harga kita lebih murah", config.COLOR_GREEN),
        ("Harga sama", config.COLOR_YELLOW),
        ("Harga kita lebih mahal", config.COLOR_RED),
        ("Tidak ada data marketplace", config.COLOR_GREY),
    ]
    for label, color in legend:
        r += 1
        c = ws.cell(row=r, column=1, value=label)
        c.fill = PatternFill("solid", fgColor=color)
        c.border = border


def write_report(df: pd.DataFrame, summary: dict, output_path: Path | None = None) -> Path:
    """Tulis Main Data + Summary ke satu file Excel dengan formatting lengkap."""
    output_path = output_path or (config.OUTPUT_DIR / config.OUTPUT_FILENAME)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    log.info("Generating Excel...")

    # Tulis Main Data dulu via pandas (cepat & aman untuk encoding).
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Main Data", index=False)
    except PermissionError:
        # Penyebab paling umum: file output sedang dibuka di Excel.
        raise PermissionError(
            f"File '{output_path.name}' sedang terbuka/terkunci (biasanya karena "
            f"masih dibuka di Excel). Tutup file tersebut lalu jalankan ulang, "
            f"atau pakai --output <nama lain>."
        ) from None

    # Buka kembali untuk formatting.
    wb = load_workbook(output_path)
    ws = wb["Main Data"]
    nrows = len(df)
    ncols = df.shape[1]

    _style_header(ws, ncols)
    _apply_number_formats(ws, df, nrows)
    _apply_row_highlight(ws, df, nrows, ncols)   # warna statis seluruh baris
    _apply_conditional_formatting(ws, df, nrows)
    ws.freeze_panes = "A2"                       # Freeze header
    ws.auto_filter.ref = ws.dimensions           # Auto Filter
    _autofit(ws, df)

    _write_summary_sheet(wb, summary)

    wb.save(output_path)
    log.info("  Tersimpan: %s (%d baris)", output_path, nrows)
    return output_path
